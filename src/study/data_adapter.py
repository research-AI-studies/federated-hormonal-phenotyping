"""Adapter from the released cohort schema to the analysis feature space.

Defines the ten hormonal/lifestyle clustering inputs and the eight EORTC
outcome scales, constructs the analytic cohort (valid diagnosis with observed
age and body mass index), winsorises implausible values, recodes contraceptive
method into four clinical bins, and exposes both the encoded clustering matrix
and the raw outcome frame.
"""
from __future__ import annotations

from dataclasses import dataclass

import numpy as np
import pandas as pd

# Ten primary clustering inputs (codebook field names).
CLUSTER_INPUTS = [
    "age", "bmi", "menstruation_firsttime_age", "pregnancy_number",
    "bust", "cupsize", "alcohol", "smokingstatus", "menopause_yn",
    "contraceptive_kind",
]
CONTINUOUS = ["age", "bmi", "menstruation_firsttime_age"]
ORDINAL = ["pregnancy_number", "bust", "cupsize", "alcohol", "smokingstatus"]
NOMINAL = ["menopause_yn", "contraceptive_kind"]

# Eight EORTC outcome scales used for phenotype-to-symptom mapping.
EORTC_SCALES = ["fa", "sl", "ef", "brbi", "brsef", "dy", "brsee", "brhl"]
EORTC_LABEL = {
    "fa": "Fatigue", "sl": "Insomnia", "ef": "Emotional functioning",
    "brbi": "Body image", "brsef": "Sexual functioning", "dy": "Dyspnoea",
    "brsee": "Sexual enjoyment", "brhl": "Hair-loss distress",
}

DIAGNOSIS_LABEL = {1: "Breast cancer", 2: "DCIS", 3: "Fibroadenoma", 4: "Other benign"}


def _recode_contraceptive(v):
    """Collapse nine raw contraceptive codes to four clinical bins (+unknown)."""
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return np.nan
    v = int(v)
    if v == 999:
        return np.nan
    if v == 0:
        return 0           # none
    if v in (1,):
        return 1           # hormonal-systemic (oral)
    if v in (2, 6):
        return 1           # hormonal-systemic (injection, implant)
    if v in (3, 5):
        return 2           # hormonal-local (ring, hormonal IUD)
    if v in (4,):
        return 1           # patch -> hormonal-systemic
    if v in (7,):
        return 3           # non-hormonal mechanical (copper IUD)
    return 4               # 888/other -> unknown auxiliary


@dataclass
class StudyData:
    raw: pd.DataFrame                 # full registry
    analytic: pd.DataFrame            # analytic cohort rows
    inputs: pd.DataFrame              # cleaned clustering inputs (analytic)
    outcomes: pd.DataFrame            # EORTC scales (analytic)
    diagnosis: pd.Series              # diagnosis code (analytic)


def _read_excel(path: str) -> pd.DataFrame:
    """Robust Excel reader (openpyxl) avoiding pandas engine edge cases."""
    import openpyxl

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    it = ws.iter_rows(values_only=True)
    header = list(next(it))
    rows = [list(r) for r in it]
    wb.close()
    return pd.DataFrame(rows, columns=header)


def load_study(path: str) -> StudyData:
    raw = _read_excel(path)

    diag = raw["diagnosis"]
    valid = diag.isin([1, 2, 3, 4])
    analytic_mask = valid & raw["age"].notna() & raw["bmi"].notna()
    a = raw[analytic_mask].copy()

    inp = a[CLUSTER_INPUTS].copy()
    # Winsorise menarche to plausible [8, 18]; out-of-range -> NaN for imputation.
    m = inp["menstruation_firsttime_age"]
    inp["menstruation_firsttime_age"] = m.where((m >= 8) & (m <= 18), np.nan)
    # Winsorise BMI to 1st/99th percentile.
    lo, hi = inp["bmi"].quantile([0.01, 0.99])
    inp["bmi"] = inp["bmi"].clip(lo, hi)
    # Recode contraceptive.
    inp["contraceptive_kind"] = inp["contraceptive_kind"].map(_recode_contraceptive)

    out = a[EORTC_SCALES].copy()

    return StudyData(
        raw=raw,
        analytic=a,
        inputs=inp,
        outcomes=out,
        diagnosis=a["diagnosis"].astype(int),
    )
